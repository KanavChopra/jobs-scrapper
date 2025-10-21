import os
import pandas as pd
import pytz
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def save_to_excel(df: pd.DataFrame, file_name: str) -> None:
    """Saves the DataFrame to an Excel file.

    Args:
        df (pd.DataFrame): The DataFrame to save.
        file_path (str): The path to the Excel file.
    """

    # Get the project root directory (where the Python file is located)
    project_root = Path(__file__).resolve().parent.parent
    output_dir = project_root / "output"

    # Create folder if missing
    output_dir.mkdir(parents=True, exist_ok=True)

    # Define the full file path
    file_path = output_dir / file_name

    # Get current date in Indian timezone for sheet name
    tz = pytz.timezone("Asia/Kolkata")
    current_date = datetime.now(tz).strftime("%Y-%m-%d")

    # Prepare the DataFrame for saving
    data = df.copy()
    data = data.rename(
        columns = {
            "companyName": "Company Name",
            "title": "Job Title",
            "location": "Location",
            "link": "Job Link",
            "applyURL": "Apply Link",
            "applicantsCount": "Applicants Count",
            "postedAt": "Posted At"
        }
    )

    # Only include columns that actually exist in the DataFrame
    expected_cols = [
        "Company Name",
        "Job Title",
        "Location",
        "Posted At",
        "Job Link",
        "Apply Link",
        "Applicants Count"
    ]

    # Filter columns
    data = data[[col for col in expected_cols if col in data.columns]]
    
    # Add hyperlinks
    if 'Job Link' in data.columns:
        data['Job Link'] = data['Job Link'].apply(lambda x: f'=HYPERLINK("{x}", "Link")' if pd.notna(x) else "")
    if 'Apply Link' in data.columns:
        data['Apply Link'] = data['Apply Link'].apply(lambda x: f'=HYPERLINK("{x}", "Apply Here")' if pd.notna(x) else "")

    # Sort by Applicants Count
    data['Applicants Count'] = pd.to_numeric(
        data['Applicants Count'].str.replace(',', ''), 
        errors='coerce'
    ).fillna(0).astype(int)
    data.sort_values(by='Applicants Count', ascending=True, inplace=True)

    # Save or append the data to Excel file
    if file_path.exists():
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="new") as writer:
            data.to_excel(writer, sheet_name=current_date, index=False)
        print(f"📘 Added new sheet '{current_date}' to existing file: {file_path}")
    else:
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            data.to_excel(writer, sheet_name=current_date, index=False)
        print(f"🆕 Created new file: {file_path} (sheet: '{current_date}')")

    # Load the workbook and add a table to the new sheet
    wb = load_workbook(file_path)
    ws = wb[current_date]

    # Define a table range
    last_row = ws.max_row
    last_col = ws.max_column

    # Create a table
    table_ref = f"A1:{chr(64 + last_col)}{last_row}"
    table_name = f"Table_{current_date.replace('-', '_')}"
    table = Table(displayName=table_name, ref=table_ref)

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style

    # Apply hyperlink font style to "Job Link" and "Apply Link" columns
    blue_link_font = Font(color="0000FF", underline="single")
    header_cells = ws[1]
    for cell in header_cells:
        if cell.value in ["Job Link", "Apply Link"]:
            col_letter = get_column_letter(cell.column)
            for row in range(2, last_row + 1):
                ws[f"{col_letter}{row}"].font = blue_link_font

    # Add the table to the worksheet
    ws.add_table(table)
    wb.save(file_path)
    wb.close()
    print(f"✅ Excel table '{table_name}' created in sheet '{current_date}'")